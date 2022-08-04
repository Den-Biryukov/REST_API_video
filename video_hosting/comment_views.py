from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_headers
from openpyxl import Workbook
from openpyxl.styles import Alignment

from video_hosting.models import *
from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import *

from django.core.exceptions import ObjectDoesNotExist
from rest_framework import status

from rest_api_video.celery import clear_comments


class CommentView(APIView):

    def post(self, request):
        video_id = request.data.get('video')
        content = request.data.get('content')
        user_id = request.data.get('user_id')
        user = User.objects.get(id=user_id)
        video = Video.objects.get(id=video_id)
        comment = Comment.objects.create(video=video, content=content)
        comment_sirialized = CommentsSerializer(comment).data
        return Response(comment_sirialized)

    @method_decorator(cache_page(60 * 2))
    @method_decorator(vary_on_headers("Authorization", ))
    def get(self, request, pk=None):
        try:
            if not pk:
                comments = Comment.objects.all()
                comment_serialized = CommentsSerializer(comments, many=True).data
                clear_comments.delay()
                return Response(comment_serialized)
            comment = Comment.objects.get(id=pk)
            comment_serialized = CommentsSerializer(comment).data
            return Response(comment_serialized)
        except ObjectDoesNotExist as e:
            return Response({'message': 'oops! 404 <>_<>.This comment does not exist'},
                            status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, *args, **kwargs):
        pk = kwargs.get('pk', None)
        if not pk:
            return Response({'error': 'Method DELETE not allowed'})

        try:
            instance = Comment.objects.get(id=pk)
            instance.delete()
        except:
            return Response({'error': 'Object not found'})

        return Response({'post': 'delete comment ' + str(pk)})


class CommentPutView(APIView):
    def put(self, request, *args, **kwargs):
        pk = kwargs.get('pk', None)
        if not pk:
            return Response({'error': 'Method PUT not allowed'})

        try:
            instance = Comment.objects.get(pk=pk)
        except:
            return Response({'error': "Object doesn't exists"})

        serializer = CommentsPutSerializer(data=request.data, instance=instance) # когда прописано два аргумента data и instance, у сериалайзера автоматически вызовется метод update, который прописан в serializes.py
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'post': serializer.data})


class ExportMyCommentsView(APIView):

    def get(self, request):
        current_user = request.user
        # my_comments = Comment.objects.filter(owner_id=current_user.id)
        my_comments = Comment.objects.all()
        alignment = Alignment(horizontal='center', vertical='center', )
        workbook = Workbook()
        ws = workbook.active
        fields = ['owner', 'video', 'content', 'Likes_count']
        column = 1
        row = 1
        for field in fields:
            cell = ws.cell(column=column, row=1, value=field)
            cell.alignment = alignment
            column += 1
        for comment in my_comments:
            row += 1
            ws.cell(column=1, row=row, value=comment.owner_id)
            ws.cell(column=2, row=row, value=comment.video_id)
            ws.cell(column=3, row=row, value=comment.content)
            ws.cell(column=4, row=row, value=comment.likes_count)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename=Data.xlsx'
        workbook.save(response)
        return response